from bs4 import BeautifulSoup
import openpyxl
import requests
from datetime import datetime
from colorama import init, Fore, Style

start_time = datetime.now()

init()
GREEN = Fore.GREEN
RED = Fore.RED
YELLOW = Fore.YELLOW
BLUE = Fore.BLUE
RESET = Style.RESET_ALL

filename = "data.xlsx"
wb = openpyxl.load_workbook(filename)
sheet = wb['Arkusz1']
rows = len(sheet['A'])
print(GREEN + '[+] Ilosc adresow url: ' + str(rows) + RESET)
counter = 2

while True:
    url = sheet['A{}'.format(counter)]
    cms = sheet['K{}'.format(counter)]
    if url.value == None:
        print(GREEN + 'Koniec pliku' + RESET)
        break
    else:
        try:

            print()
            print("---------------------------")
            print(url.value)
            print(GREEN + "[+] Adres " + str(counter) + " z " + str(rows) + RESET)
            print("---------------------------")
            print()
            
            headers = {
    "User-Agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    }
            r = requests.get(url.value, headers=headers, timeout=10)
            soup = BeautifulSoup(r.text, 'html.parser')
            shoper = soup.find('div', {'id' :'shoper-foot'})
            wix = soup.find('script', {'type' : 'application/json', 'id' :'wix-warmup-data'})
            joomla = soup.find('meta', {'name' : 'generator'})       
            shopgold = soup.find('div', {"class" : "Copy"})

            if shoper:
                print(BLUE + url.value + " SHOPER" + RESET)
                sheet.cell(row=counter, column=10).value = 'shoper'
                wb.save(filename)
            elif wix:
                print(BLUE + url.value + " WIX" + RESET)
                sheet.cell(row=counter, column=10).value = 'wix'
                wb.save(filename)
            elif shopgold != None:
                if 'shopGold' in shopgold.text:
                    print(BLUE + url.value + " SHOPGOLD" + RESET)
                    sheet.cell(row=counter, column=10).value = 'shopgold'
                    wb.save(filename)
            elif joomla != None:
                if 'Joomla' in joomla['content']:
                    print(BLUE + url.value + " JOOMLA" + RESET)
                    sheet.cell(row=counter, column=10).value = 'joomla'
                    wb.save(filename)            
            else:
                pass
            
        except Exception as e:
            sheet.cell(row=counter, column=10).value = 'blad polaczenia'
            wb.save(filename)           
            print(RED + str(e) + RESET)

        except KeyboardInterrupt:
            print(RED + '[X] Koncze dzialanie' + RESET)
            quit()
        counter += 1

end_time = datetime.now()

print(YELLOW + 'Czas trwania: {}'.format(end_time - start_time) + RESET)



